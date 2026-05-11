from __future__ import annotations

import os
import re
import sys
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
QUOTE_APP_DIR = ROOT_DIR / "quote_app"
if str(QUOTE_APP_DIR) not in sys.path:
    sys.path.insert(0, str(QUOTE_APP_DIR))

from streamlit.testing.v1 import AppTest


CONTENT_CONFIRMATION_RE = re.compile(r"需确认|待确认|人工确认|确认|核价|核定|报价口径|计价|测算")


SAMPLE_TEXT = """
活动内容规划
（一）板块一：琼韵启幕礼
设置启动仪式，安排主持人、领导致辞。

（二）板块二：滨海非遗生活市集
设置30个市集摊位，配置导视指引。

（三）板块三：海风音乐会
安排民乐演出和青年乐队演唱。

宣传排期
倒计时海报
公众号软文

活动保障
医疗保障
交通引导
"""


def _new_app() -> AppTest:
    os.chdir(QUOTE_APP_DIR)
    app = AppTest.from_file("app.py", default_timeout=20)
    app.run()
    return app


def _click_button(app: AppTest, label: str) -> None:
    for button in app.button:
        if button.label == label:
            button.click()
            return
    raise AssertionError(f"button not found: {label}; buttons={[button.label for button in app.button]}")


def _recognize_sample() -> AppTest:
    app = _new_app()
    app.text_area[0].set_value(SAMPLE_TEXT)
    _click_button(app, "识别报价项")
    app.run()
    if app.exception:
        raise AssertionError(f"app exceptions: {app.exception}")
    return app


def _assert_no_confirmation_content(df: pd.DataFrame, label: str) -> None:
    if df.empty or "内容/尺寸/工艺" not in df.columns:
        return
    bad_rows = df[df["内容/尺寸/工艺"].astype(str).str.contains(CONTENT_CONFIRMATION_RE, na=False)]
    if not bad_rows.empty:
        display_cols = [column for column in ["项目分类", "项目", "标准项目", "内容/尺寸/工艺"] if column in bad_rows.columns]
        raise AssertionError(f"{label} 内容/尺寸/工艺 contains confirmation wording: {bad_rows[display_cols].to_dict('records')}")


def test_review_edit_requires_submit() -> None:
    app = _recognize_sample()
    before_keep = bool(app.session_state["working_quote_df"].iloc[0]["是否保留"])
    app.session_state["review_quote_editor"] = {"edited_rows": {0: {"是否加入": False}}, "added_rows": [], "deleted_rows": []}
    app.run()
    after_plain_run_keep = bool(app.session_state["working_quote_df"].iloc[0]["是否保留"])
    if before_keep is not True or after_plain_run_keep is not True:
        raise AssertionError("review edits should not apply before clicking 应用本表修改")


def test_review_remove_updates_final_quote() -> None:
    app = _recognize_sample()
    before_count = len(app.session_state["final_quote_df"])
    app.session_state["review_quote_editor"] = {"edited_rows": {0: {"是否加入": False}}, "added_rows": [], "deleted_rows": []}
    _click_button(app, "应用本表修改")
    app.run()
    after_count = len(app.session_state["final_quote_df"])
    if bool(app.session_state["working_quote_df"].iloc[0]["是否保留"]):
        raise AssertionError("review removal did not update working_quote_df")
    if after_count >= before_count:
        raise AssertionError(f"final quote did not shrink after removal: before={before_count}, after={after_count}")


def test_final_edit_and_export() -> None:
    app = _recognize_sample()
    _assert_no_confirmation_content(app.session_state["working_quote_df"], "working_quote_df")
    _assert_no_confirmation_content(app.session_state["final_quote_df"], "final_quote_df")
    app.session_state["final_quote_editor"] = {"edited_rows": {0: {"数量": 2.0, "单价": 10.0}}, "added_rows": [], "deleted_rows": []}
    _click_button(app, "保存报价单编辑")
    app.run()
    row = app.session_state["final_quote_df"].iloc[0]
    if float(row["数量"]) != 2.0 or float(row["单价"]) != 10.0 or float(row["合计"]) != 20.0:
        raise AssertionError(f"final edit not applied: {row.to_dict()}")

    _click_button(app, "生成 Excel 报价单")
    app.run()
    excel_bytes = app.session_state["excel_bytes"]
    if not isinstance(excel_bytes, bytes) or len(excel_bytes) < 1000:
        raise AssertionError("Excel export did not produce bytes")
    workbook = load_workbook(BytesIO(excel_bytes), data_only=False)
    sheet = workbook.active
    bad_values = [
        str(sheet.cell(row=row, column=4).value)
        for row in range(4, sheet.max_row)
        if sheet.cell(row=row, column=4).value and CONTENT_CONFIRMATION_RE.search(str(sheet.cell(row=row, column=4).value))
    ]
    if bad_values:
        raise AssertionError(f"Excel 内容/尺寸/工艺 contains confirmation wording: {bad_values}")


def test_quantity_and_public_support_sections() -> None:
    app = _recognize_sample()
    working_df = app.session_state["working_quote_df"]
    stall_rows = working_df[working_df["标准项目"].astype(str) == "帐篷摊位"]
    if stall_rows.empty:
        raise AssertionError("missing 帐篷摊位 row")
    quantity = float(stall_rows.iloc[0]["数量"])
    if quantity != 30.0:
        raise AssertionError(f"市集摊位数量应为 30，不应因多别名命中翻倍：{quantity}")

    support_rows = working_df[working_df["标准项目"].astype(str).isin(["医疗保障", "交通停车引导"])]
    if support_rows.empty:
        raise AssertionError("missing support rows")
    bad_sections = support_rows[support_rows["quote_section"].astype(str) != "人员类及其他"]
    if not bad_sections.empty:
        raise AssertionError(f"保障项不应继承最后一个活动板块：{bad_sections[['标准项目', 'quote_section']].to_dict('records')}")


def test_suggestion_edit_requires_submit() -> None:
    app = _recognize_sample()
    if app.session_state["suggestion_editor_df"].empty:
        raise AssertionError("sample should produce suggestion rows")
    _assert_no_confirmation_content(app.session_state["suggestion_editor_df"], "suggestion_editor_df")

    before_count = len(app.session_state["final_quote_df"])
    app.session_state["suggestion_quote_editor"] = {"edited_rows": {0: {"是否加入报价单": True}}, "added_rows": [], "deleted_rows": []}
    app.run()
    after_plain_run_count = len(app.session_state["final_quote_df"])
    if after_plain_run_count != before_count:
        raise AssertionError("suggestion edits should not apply before clicking 应用建议补充项")


def test_candidate_feedback_edit_requires_submit() -> None:
    app = _recognize_sample()
    candidate_rows = pd.DataFrame(
        [
            {
                "是否处理": False,
                "候选词": "星空泡泡互动点",
                "可能归类": "趣味互动游戏",
                "上下文摘要": "设置星空泡泡互动点",
                "处理方式": "暂不处理",
                "选择标准项目": "",
                "备注": "",
            }
        ]
    )
    app.session_state["candidate_display_df"] = candidate_rows.copy()
    app.session_state["feedback_df"] = candidate_rows.copy()
    app.run()

    app.session_state["candidate_feedback_editor"] = {"edited_rows": {0: {"是否处理": True}}, "added_rows": [], "deleted_rows": []}
    app.run()
    if bool(app.session_state["feedback_df"].iloc[0]["是否处理"]):
        raise AssertionError("candidate feedback edits should not apply before clicking 保存规则更新")


def main() -> None:
    tests = [
        test_review_edit_requires_submit,
        test_review_remove_updates_final_quote,
        test_final_edit_and_export,
        test_quantity_and_public_support_sections,
        test_suggestion_edit_requires_submit,
        test_candidate_feedback_edit_requires_submit,
    ]
    for test in tests:
        test()
        print(f"PASS {test.__name__}")
    print("ALL APP FLOW TESTS PASSED")


if __name__ == "__main__":
    main()
