"""Export quote rows to a styled Excel workbook."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .quote_builder import dedupe_final_quote_items, sort_by_quote_section


EXPORT_HEADERS = ["序号", "项目分类", "项目", "内容/尺寸/工艺", "数量", "单位", "单价", "合计（元）", "备注"]


def _to_number(value: Any) -> float | None:
    if pd.isna(value) or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _display_value(value: Any) -> Any:
    if pd.isna(value):
        return ""
    return value


def export_quote_to_excel(
    quote_df: pd.DataFrame,
    activity_name: str,
    activity_time: str,
    client_name: str,
    output_path: str | Path | None = None,
) -> bytes:
    """Create a formal quote workbook and return its bytes."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报价单"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title = f"{activity_name or '活动'}报价单"

    sheet.merge_cells("A1:I1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet["A2"] = "活动时间："
    sheet["B2"] = activity_time
    sheet["F2"] = "单位："
    sheet["G2"] = client_name

    for col_index, header in enumerate(EXPORT_HEADERS, start=1):
        cell = sheet.cell(row=3, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    kept_df = sort_by_quote_section(dedupe_final_quote_items(quote_df.copy()))
    if "是否保留" in kept_df.columns:
        kept_df = kept_df[kept_df["是否保留"].astype(bool)]
    kept_df = sort_by_quote_section(kept_df)

    start_row = 4
    for offset, (_, row) in enumerate(kept_df.iterrows(), start=0):
        excel_row = start_row + offset
        quantity = _to_number(row.get("数量"))
        unit_price = _to_number(row.get("单价"))
        total = 0 if quantity is None or unit_price is None else quantity * unit_price

        values = [
            offset + 1,
            _display_value(row.get("项目分类")),
            _display_value(row.get("标准项目", row.get("项目"))),
            _display_value(row.get("内容/尺寸/工艺")),
            quantity if quantity is not None else _display_value(row.get("数量")),
            _display_value(row.get("单位")),
            unit_price if unit_price is not None else "",
            total,
            _display_value(row.get("备注")),
        ]

        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=excel_row, column=col_index, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    last_detail_row = start_row + len(kept_df) - 1
    total_row = start_row + len(kept_df)
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    sheet.cell(row=total_row, column=1, value="合计")
    sheet.cell(row=total_row, column=8, value=f"=SUM(H{start_row}:H{last_detail_row})" if len(kept_df) else 0)
    sheet.cell(row=total_row, column=1).font = Font(bold=True)
    sheet.cell(row=total_row, column=8).font = Font(bold=True)

    max_row = total_row
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            if cell.row != 1:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(4, max_row + 1):
        sheet.cell(row=row, column=7).number_format = '#,##0.00'
        sheet.cell(row=row, column=8).number_format = '#,##0.00'

    for column in range(1, 10):
        letter = get_column_letter(column)
        max_length = 0
        for cell in sheet[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[letter].width = min(max(max_length + 4, 10), 35)

    sheet.column_dimensions["D"].width = 38
    sheet.column_dimensions["I"].width = 42
    sheet.freeze_panes = "A4"

    buffer = BytesIO()
    workbook.save(buffer)
    data = buffer.getvalue()

    if output_path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(data)

    return data
